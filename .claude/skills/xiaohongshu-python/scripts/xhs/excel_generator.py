"""Excel 潜客管理表生成器 — openpyxl 实现。

行粒度: 一行 = 一个用户在一个帖子下的所有评论（合并）。
同一用户在不同帖子各占一行，方便筛选。

用法（通过 cli.py）:
    python scripts/cli.py generate-excel --input data/analysis.json
    python scripts/cli.py generate-excel --input data/analysis.json --output custom.xlsx
"""

from __future__ import annotations

import json
import os
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─── 常量 ───

CIRCLE_NUMS = ["①", "②", "③", "④", "⑤", "⑥", "⑦", "⑧", "⑨", "⑩"]

# 16 列定义
COLUMNS = [
    # 用户信息区（蓝色）
    {"header": "用户名", "key": "username", "width": 18},
    {"header": "用户主页", "key": "profileUrl", "width": 20},
    {"header": "IP属地", "key": "ipLocation", "width": 10},
    # 兴趣分析区（绿色）
    {"header": "评论数量", "key": "commentCount", "width": 10},
    {"header": "评论内容", "key": "content", "width": 55},
    {"header": "兴趣得分", "key": "interestScore", "width": 10},
    {"header": "兴趣标签", "key": "interestTags", "width": 25},
    {"header": "判断理由", "key": "reason", "width": 45},
    # 来源帖子区（金色）
    {"header": "帖子标题", "key": "postTitle", "width": 40},
    {"header": "帖子链接", "key": "postUrl", "width": 20},
    {"header": "帖子截图", "key": "screenshot", "width": 25},
    {"header": "帖子总评论数", "key": "totalComments", "width": 14},
    {"header": "本次获取评论数", "key": "collectedComments", "width": 16},
    # 跟进管理区（红色）
    {"header": "已关注", "key": "followed", "width": 10},
    {"header": "跟进状态", "key": "followUpStatus", "width": 14},
    {"header": "负责人", "key": "owner", "width": 12},
]

# 4 个区域的表头颜色（ARGB 格式）
HEADER_COLORS = {
    "user": "FF2E75B6",       # 蓝色
    "interest": "FF548235",   # 绿色
    "post": "FFBF8F00",       # 金色
    "followUp": "FFC00000",   # 红色
}

# 每列对应哪个区域
HEADER_COLOR_MAP = [
    "user", "user", "user",
    "interest", "interest", "interest", "interest", "interest",
    "post", "post", "post", "post", "post",
    "followUp", "followUp", "followUp",
]

# 交替行背景色
BAND_COLORS = ["FFF2F7FB", "FFFFFFFF"]

# 得分高亮
SCORE_HIGH_FILL = PatternFill("solid", fgColor="FFC6EFCE")
SCORE_HIGH_FONT_COLOR = "FF006100"
SCORE_MID_FILL = PatternFill("solid", fgColor="FFFFEB9C")
SCORE_MID_FONT_COLOR = "FF9C5700"

# 边框
THIN_BORDER_BOTTOM = Border(bottom=Side(style="thin", color="FFD9D9D9"))

# 下拉选项
FOLLOWED_OPTIONS = "是,否"
FOLLOW_UP_OPTIONS = "待跟进,已联系,有意向,已成交,已流失"


# ─── 工具函数 ───


def _group_comments_by_user(comments: list[dict]) -> list[dict]:
    """将同一帖子下同一用户的多条评论合并为一条记录。"""
    user_map: dict[str, dict] = {}
    order: list[str] = []

    for c in comments:
        key = c.get("userId") or c.get("username", "")
        if not key:
            continue
        if key not in user_map:
            user_map[key] = {
                "username": c.get("username", ""),
                "userId": c.get("userId", ""),
                "profileUrl": c.get("profileUrl", ""),
                "ipLocation": c.get("ipLocation", ""),
                "contents": [],
                "scores": [],
                "tags": set(),
                "reasons": [],
            }
            order.append(key)
        user = user_map[key]
        user["contents"].append(c.get("content", ""))
        user["scores"].append(c.get("interestScore", 0))
        if c.get("interestTags"):
            for tag in str(c["interestTags"]).replace("，", ",").split(","):
                tag = tag.strip()
                if tag:
                    user["tags"].add(tag)
        if c.get("reason"):
            user["reasons"].append(c["reason"])
        if c.get("ipLocation") and not user["ipLocation"]:
            user["ipLocation"] = c["ipLocation"]

    result = []
    for key in order:
        user = user_map[key]
        count = len(user["contents"])

        if count == 1:
            merged_content = user["contents"][0]
        else:
            merged_content = "\n".join(
                f"{CIRCLE_NUMS[i] if i < len(CIRCLE_NUMS) else f'({i + 1})'}{text}"
                for i, text in enumerate(user["contents"])
            )

        max_score = max(user["scores"]) if user["scores"] else 0

        unique_reasons = list(dict.fromkeys(user["reasons"]))
        if len(unique_reasons) <= 1:
            merged_reason = unique_reasons[0] if unique_reasons else ""
        else:
            merged_reason = "\n".join(
                f"{CIRCLE_NUMS[i] if i < len(CIRCLE_NUMS) else f'({i + 1})'}{r}"
                for i, r in enumerate(unique_reasons)
            )

        result.append({
            "username": user["username"],
            "profileUrl": user["profileUrl"],
            "ipLocation": user["ipLocation"],
            "commentCount": count,
            "content": merged_content,
            "interestScore": max_score,
            "interestTags": ", ".join(sorted(user["tags"])),
            "reason": merged_reason,
        })

    result.sort(key=lambda x: x["interestScore"], reverse=True)
    return result


def generate_excel(input_path: str, output_path: str | None = None) -> str:
    """从 analysis.json 生成 Excel 潜客管理表。

    返回生成的文件路径。
    """
    with open(input_path, encoding="utf-8") as f:
        data = json.load(f)

    if not output_path:
        keyword = data.get("keyword", "export")
        date_str = datetime.now(tz=UTC).strftime("%Y%m%d")
        script_dir = Path(__file__).resolve().parent.parent.parent
        output_dir = script_dir / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = str(output_dir / f"xhs-{keyword}-{date_str}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "潜客管理"

    # 列宽
    for i, col_def in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = col_def["width"]

    # 表头行
    for i, col_def in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i, value=col_def["header"])
        color_key = HEADER_COLOR_MAP[i - 1]
        cell.font = Font(bold=True, color="FFFFFFFF", size=11, name="Arial")
        cell.fill = PatternFill("solid", fgColor=HEADER_COLORS[color_key])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # 数据验证
    dv_followed = DataValidation(type="list", formula1=f'"{FOLLOWED_OPTIONS}"', allow_blank=True)
    dv_followup = DataValidation(type="list", formula1=f'"{FOLLOW_UP_OPTIONS}"', allow_blank=True)
    ws.add_data_validation(dv_followed)
    ws.add_data_validation(dv_followup)

    current_row = 2
    total_users = 0
    added_images: dict[str, XlImage] = {}

    link_font = Font(size=10, name="Arial", color="FF0563C1", underline="single")
    default_font = Font(size=10, name="Arial")

    for pi, post in enumerate(data.get("posts", [])):
        comments = post.get("validComments", [])
        collected_comments = post.get("collectedComments", len(comments))
        band_color = BAND_COLORS[pi % 2]
        band_fill = PatternFill("solid", fgColor=band_color)

        user_rows = _group_comments_by_user(comments)
        if not user_rows:
            continue

        start_row = current_row

        for user_row in user_rows:
            row_num = current_row

            # 用户信息区
            ws.cell(row=row_num, column=1, value=user_row["username"])
            if user_row["profileUrl"]:
                cell_b = ws.cell(row=row_num, column=2)
                cell_b.value = user_row["profileUrl"]
                cell_b.hyperlink = user_row["profileUrl"]
                cell_b.font = link_font
            ws.cell(row=row_num, column=3, value=user_row["ipLocation"])

            # 兴趣分析区
            ws.cell(row=row_num, column=4, value=user_row["commentCount"])
            ws.cell(row=row_num, column=5, value=user_row["content"])
            ws.cell(row=row_num, column=6, value=user_row["interestScore"])
            ws.cell(row=row_num, column=7, value=user_row["interestTags"])
            ws.cell(row=row_num, column=8, value=user_row["reason"])

            # 来源帖子区
            ws.cell(row=row_num, column=9, value=post.get("title", ""))
            if post.get("url"):
                cell_j = ws.cell(row=row_num, column=10)
                cell_j.value = post["url"]
                cell_j.hyperlink = post["url"]
                cell_j.font = link_font
            ws.cell(row=row_num, column=11, value="")  # 截图占位
            ws.cell(row=row_num, column=12, value=post.get("totalComments", ""))
            ws.cell(row=row_num, column=13, value=collected_comments)

            # 跟进管理区（空）
            ws.cell(row=row_num, column=14, value="")
            ws.cell(row=row_num, column=15, value="")
            ws.cell(row=row_num, column=16, value="")

            # 行高
            line_count = user_row["content"].count("\n") + 1
            ws.row_dimensions[row_num].height = max(25, line_count * 18)

            # 行样式
            for col in range(1, 17):
                cell = ws.cell(row=row_num, column=col)
                if not cell.font or cell.font == Font():
                    cell.font = default_font
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.fill = band_fill
                cell.border = THIN_BORDER_BOTTOM

            # 评论数量居中
            ws.cell(row=row_num, column=4).alignment = Alignment(
                horizontal="center", vertical="center"
            )

            # 得分高亮
            score_cell = ws.cell(row=row_num, column=6)
            score_cell.alignment = Alignment(horizontal="center", vertical="center")
            score_cell.font = Font(bold=True, size=10, name="Arial")
            score = user_row["interestScore"]
            if score >= 8:
                score_cell.fill = SCORE_HIGH_FILL
                score_cell.font = Font(
                    bold=True, size=10, name="Arial", color=SCORE_HIGH_FONT_COLOR
                )
            elif score >= 6:
                score_cell.fill = SCORE_MID_FILL
                score_cell.font = Font(
                    bold=True, size=10, name="Arial", color=SCORE_MID_FONT_COLOR
                )

            # 居中列
            for col in (12, 13, 14, 15):
                ws.cell(row=row_num, column=col).alignment = Alignment(
                    horizontal="center", vertical="center"
                )

            # 数据验证
            dv_followed.add(ws.cell(row=row_num, column=14))
            dv_followup.add(ws.cell(row=row_num, column=15))

            current_row += 1
            total_users += 1

        # 嵌入帖子截图
        screenshot = post.get("screenshotFile", "")
        if screenshot and os.path.isfile(screenshot):
            if screenshot not in added_images:
                img = XlImage(screenshot)
                img.width = 150
                img.height = 80
                added_images[screenshot] = img

            img = added_images[screenshot]
            anchor = f"K{start_row}"
            ws.add_image(img, anchor)
            ws.row_dimensions[start_row].height = max(
                ws.row_dimensions[start_row].height or 25, 80
            )

    # 自动筛选
    if current_row > 2:
        ws.auto_filter.ref = f"A1:P{current_row - 1}"

    # 保存
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)

    return output_path, len(data.get("posts", [])), total_users, len(added_images)
